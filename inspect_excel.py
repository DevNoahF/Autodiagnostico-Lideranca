from openpyxl import load_workbook
from pathlib import Path

p = Path('base/PTT_Autodiagnostico_Lideranca_GC.xlsx')
wb = load_workbook(p, data_only=True)
print('SHEETS:', wb.sheetnames)
for ws in wb.worksheets:
    print('--- SHEET:', ws.title)
    print('ROWS:', ws.max_row, 'COLS:', ws.max_column)
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True):
        print(row)
